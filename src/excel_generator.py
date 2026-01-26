"""
Excel Workbook Generator

Creates downloadable Excel workbooks with comprehensive company analysis.
This is what analysts actually want - structured data they can work with.

Tabs:
1. Company Overview - Financials, management, key metrics
2. Pipeline Summary - All assets with phase/indication
3. Clinical Data - Efficacy/safety comparisons
4. Ownership - 13F holders with QoQ changes
5. Catalysts - Upcoming events and dates
6. KOLs - Key opinion leaders with contact info
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from pathlib import Path
from dataclasses import dataclass
from typing import Optional
import json


# Style definitions
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


@dataclass
class CompanyData:
    """All data needed for a company workbook"""
    ticker: str
    name: str
    description: str = ""
    
    # Financials
    market_cap_mm: float = 0
    cash_mm: float = 0
    burn_rate_mm: float = 0
    runway_months: int = 0
    
    # Pipeline
    pipeline: list = None
    
    # Ownership
    ownership: list = None
    
    # Catalysts
    catalysts: list = None
    
    # KOLs
    kols: list = None
    
    # Clinical data
    clinical_data: list = None
    
    def __post_init__(self):
        self.pipeline = self.pipeline or []
        self.ownership = self.ownership or []
        self.catalysts = self.catalysts or []
        self.kols = self.kols or []
        self.clinical_data = self.clinical_data or []


class ExcelWorkbookGenerator:
    """Generates comprehensive Excel workbooks for biotech companies"""
    
    def __init__(self, output_dir: str = "data/workbooks"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate(self, company: CompanyData) -> str:
        """
        Generate a complete company analysis workbook.
        
        Returns path to the generated file.
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        self._create_overview_sheet(wb, company)
        self._create_pipeline_sheet(wb, company)
        self._create_clinical_sheet(wb, company)
        self._create_ownership_sheet(wb, company)
        self._create_catalysts_sheet(wb, company)
        self._create_kol_sheet(wb, company)
        
        filename = f"{company.ticker}_analysis_{date.today().strftime('%Y%m%d')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        
        print(f"Generated workbook: {filepath}")
        return str(filepath)
    
    def _style_header_row(self, ws, row: int, num_cols: int):
        """Apply header styling to a row"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _auto_column_width(self, ws):
        """Auto-adjust column widths based on content"""
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            length = min(length + 2, 50)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
    
    def _create_overview_sheet(self, wb: Workbook, company: CompanyData):
        """Create the company overview sheet"""
        ws = wb.create_sheet("Overview")
        
        ws.merge_cells('A1:D1')
        ws['A1'] = f"{company.ticker} - {company.name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, color="666666")
        
        row = 4
        ws.cell(row=row, column=1, value="KEY METRICS").font = Font(bold=True, size=12)
        row += 1
        
        metrics = [
            ("Market Cap", f"${company.market_cap_mm:,.0f}M"),
            ("Cash Position", f"${company.cash_mm:,.0f}M"),
            ("Quarterly Burn", f"${company.burn_rate_mm:,.0f}M"),
            ("Cash Runway", f"{company.runway_months} months"),
        ]
        
        for metric_name, metric_value in metrics:
            ws.cell(row=row, column=1, value=metric_name)
            ws.cell(row=row, column=2, value=metric_value)
            row += 1
        
        row += 1
        ws.cell(row=row, column=1, value="COMPANY DESCRIPTION").font = Font(bold=True, size=12)
        row += 1
        ws.merge_cells(f'A{row}:D{row+3}')
        ws.cell(row=row, column=1, value=company.description)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        
        self._auto_column_width(ws)
    
    def _create_pipeline_sheet(self, wb: Workbook, company: CompanyData):
        """Create the pipeline summary sheet"""
        ws = wb.create_sheet("Pipeline")
        
        headers = ["Asset Name", "Indication", "Phase", "Status", "Mechanism", "Modality", "Route", "NCT ID"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))
        
        for row, asset in enumerate(company.pipeline, 2):
            ws.cell(row=row, column=1, value=asset.get("name", ""))
            ws.cell(row=row, column=2, value=asset.get("indication", ""))
            ws.cell(row=row, column=3, value=asset.get("phase", ""))
            ws.cell(row=row, column=4, value=asset.get("status", ""))
            ws.cell(row=row, column=5, value=asset.get("mechanism", ""))
            ws.cell(row=row, column=6, value=asset.get("modality", ""))
            ws.cell(row=row, column=7, value=asset.get("route", ""))
            ws.cell(row=row, column=8, value=asset.get("nct_id", ""))
            
            phase = asset.get("phase", "").lower()
            if "3" in phase or "filed" in phase or "approved" in phase:
                ws.cell(row=row, column=3).fill = POSITIVE_FILL
            elif "2" in phase:
                ws.cell(row=row, column=3).fill = NEUTRAL_FILL
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = BORDER
        
        self._auto_column_width(ws)
    
    def _create_clinical_sheet(self, wb: Workbook, company: CompanyData):
        """Create the clinical data comparison sheet"""
        ws = wb.create_sheet("Clinical Data")
        
        headers = ["Trial / Drug", "Indication", "N", "Primary Endpoint", "Result", "p-value", "Comparator", "Comparator Result", "Source"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))
        
        for row, trial in enumerate(company.clinical_data, 2):
            ws.cell(row=row, column=1, value=trial.get("name", ""))
            ws.cell(row=row, column=2, value=trial.get("indication", ""))
            ws.cell(row=row, column=3, value=trial.get("n", ""))
            ws.cell(row=row, column=4, value=trial.get("endpoint", ""))
            ws.cell(row=row, column=5, value=trial.get("result", ""))
            ws.cell(row=row, column=6, value=trial.get("p_value", ""))
            ws.cell(row=row, column=7, value=trial.get("comparator", ""))
            ws.cell(row=row, column=8, value=trial.get("comparator_result", ""))
            ws.cell(row=row, column=9, value=trial.get("source", ""))
            
            if trial.get("met_endpoint"):
                ws.cell(row=row, column=5).fill = POSITIVE_FILL
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = BORDER
        
        if not company.clinical_data:
            ws.cell(row=2, column=1, value="No clinical data available - add trial results to populate this sheet")
        
        self._auto_column_width(ws)
    
    def _create_ownership_sheet(self, wb: Workbook, company: CompanyData):
        """Create the 13F ownership sheet"""
        ws = wb.create_sheet("Ownership (13F)")
        
        headers = ["Fund Name", "Shares", "Value ($M)", "% of Portfolio", "QoQ Change", "Change %", "Type"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))
        
        for row, holder in enumerate(company.ownership, 2):
            ws.cell(row=row, column=1, value=holder.get("fund_name", ""))
            ws.cell(row=row, column=2, value=holder.get("shares", 0))
            ws.cell(row=row, column=3, value=holder.get("value_mm", 0))
            ws.cell(row=row, column=4, value=holder.get("pct_portfolio", 0))
            ws.cell(row=row, column=5, value=holder.get("change", ""))
            ws.cell(row=row, column=6, value=holder.get("change_pct", ""))
            ws.cell(row=row, column=7, value=holder.get("fund_type", ""))
            
            # Format numbers
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3).number_format = '#,##0.0'
            ws.cell(row=row, column=4).number_format = '0.00%'
            
            # Color code changes
            change = holder.get("change", "")
            if change == "NEW" or (isinstance(holder.get("change_pct"), (int, float)) and holder.get("change_pct", 0) > 0):
                ws.cell(row=row, column=5).fill = POSITIVE_FILL
                ws.cell(row=row, column=6).fill = POSITIVE_FILL
            elif change == "CLOSED" or (isinstance(holder.get("change_pct"), (int, float)) and holder.get("change_pct", 0) < 0):
                ws.cell(row=row, column=5).fill = NEGATIVE_FILL
                ws.cell(row=row, column=6).fill = NEGATIVE_FILL
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = BORDER
        
        if not company.ownership:
            ws.cell(row=2, column=1, value="No 13F ownership data available - run 13F scraper to populate")
        
        self._auto_column_width(ws)
    
    def _create_catalysts_sheet(self, wb: Workbook, company: CompanyData):
        """Create the catalysts sheet"""
        ws = wb.create_sheet("Catalysts")
        
        headers = ["Date", "Event Type", "Asset", "Description", "Source"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))
        
        # Sort catalysts by date
        sorted_catalysts = sorted(
            company.catalysts, 
            key=lambda x: x.get("date", "9999-99-99")
        )
        
        for row, catalyst in enumerate(sorted_catalysts, 2):
            ws.cell(row=row, column=1, value=catalyst.get("date", ""))
            ws.cell(row=row, column=2, value=catalyst.get("type", ""))
            ws.cell(row=row, column=3, value=catalyst.get("asset", ""))
            ws.cell(row=row, column=4, value=catalyst.get("description", ""))
            ws.cell(row=row, column=5, value=catalyst.get("source", ""))
            
            # Highlight PDUFA dates
            if "PDUFA" in catalyst.get("type", "").upper():
                ws.cell(row=row, column=2).fill = POSITIVE_FILL
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = BORDER
        
        if not company.catalysts:
            ws.cell(row=2, column=1, value="No catalyst data available - add upcoming events to populate")
        
        self._auto_column_width(ws)
    
    def _create_kol_sheet(self, wb: Workbook, company: CompanyData):
        """Create the KOL sheet"""
        ws = wb.create_sheet("KOLs")
        
        headers = ["Name", "Email", "Institution", "Department", "Country", "Publications", "First Author", "Last Author", "Recent (3yr)"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._style_header_row(ws, 1, len(headers))
        
        for row, kol in enumerate(company.kols, 2):
            ws.cell(row=row, column=1, value=kol.get("name", ""))
            ws.cell(row=row, column=2, value=kol.get("email", ""))
            ws.cell(row=row, column=3, value=kol.get("institution", ""))
            ws.cell(row=row, column=4, value=kol.get("department", ""))
            ws.cell(row=row, column=5, value=kol.get("country", ""))
            ws.cell(row=row, column=6, value=kol.get("publication_count", 0))
            ws.cell(row=row, column=7, value=kol.get("first_author_count", 0))
            ws.cell(row=row, column=8, value=kol.get("last_author_count", 0))
            ws.cell(row=row, column=9, value=kol.get("recent_publication_count", 0))
            
            # Highlight KOLs with email addresses
            if kol.get("email"):
                ws.cell(row=row, column=2).fill = POSITIVE_FILL
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = BORDER
        
        if not company.kols:
            ws.cell(row=2, column=1, value="No KOL data available - run PubMed KOL extractor to populate")
        
        self._auto_column_width(ws)


def generate_sample_workbook():
    """Generate a sample workbook to demonstrate the format"""
    
    company = CompanyData(
        ticker="ABVX",
        name="Abivax SA",
        description="Abivax is a clinical-stage biotechnology company focused on developing therapeutics that modulate the immune system to treat inflammatory diseases.",
        market_cap_mm=850,
        cash_mm=320,
        burn_rate_mm=45,
        runway_months=21,
        pipeline=[
            {
                "name": "Obefazimod",
                "indication": "Ulcerative Colitis",
                "phase": "Phase 3 / NDA Filed",
                "status": "Active",
                "mechanism": "RNA splicing modulator",
                "modality": "Small molecule",
                "route": "Oral",
                "nct_id": "NCT05507892"
            },
            {
                "name": "Obefazimod",
                "indication": "Crohn's Disease",
                "phase": "Phase 2b",
                "status": "Enrolling",
                "mechanism": "RNA splicing modulator",
                "modality": "Small molecule",
                "route": "Oral",
                "nct_id": "NCT05678901"
            },
        ],
        ownership=[
            {"fund_name": "RA Capital Management", "shares": 2500000, "value_mm": 45.2, "pct_portfolio": 0.032, "change": "NEW", "change_pct": None, "fund_type": "Specialist"},
            {"fund_name": "Baker Bros. Advisors", "shares": 1800000, "value_mm": 32.5, "pct_portfolio": 0.018, "change": "+25%", "change_pct": 25, "fund_type": "Specialist"},
            {"fund_name": "Perceptive Advisors", "shares": 1200000, "value_mm": 21.7, "pct_portfolio": 0.015, "change": "UNCH", "change_pct": 0, "fund_type": "Specialist"},
        ],
        catalysts=[
            {"date": "2025-06-15", "type": "PDUFA", "asset": "Obefazimod", "description": "FDA decision for UC indication", "source": "Company PR"},
            {"date": "2025-03-20", "type": "Conference", "asset": "Obefazimod", "description": "DDW 2025 presentation", "source": "Conference schedule"},
            {"date": "2025-09-01", "type": "Data Readout", "asset": "Obefazimod", "description": "Crohn's Phase 2b topline", "source": "Guidance"},
        ],
        kols=[
            {"name": "William Sandborn", "email": "wsandborn@health.ucsd.edu", "institution": "UC San Diego", "department": "Gastroenterology", "country": "USA", "publication_count": 45, "first_author_count": 12, "last_author_count": 20, "recent_publication_count": 15},
            {"name": "Brian Feagan", "email": "brian.feagan@uwo.ca", "institution": "Western University", "department": "Medicine", "country": "Canada", "publication_count": 38, "first_author_count": 8, "last_author_count": 18, "recent_publication_count": 12},
        ],
        clinical_data=[
            {"name": "ABTECT-1", "indication": "Ulcerative Colitis", "n": 254, "endpoint": "Clinical Remission at Week 12", "result": "45.2%", "p_value": "<0.001", "comparator": "Placebo", "comparator_result": "12.1%", "source": "Phase 3 topline", "met_endpoint": True},
            {"name": "ABTECT-2", "indication": "Ulcerative Colitis", "n": 489, "endpoint": "Clinical Remission at Week 52", "result": "52.8%", "p_value": "<0.001", "comparator": "Placebo", "comparator_result": "18.4%", "source": "Phase 3 topline", "met_endpoint": True},
        ]
    )
    
    generator = ExcelWorkbookGenerator()
    filepath = generator.generate(company)
    
    return filepath


if __name__ == "__main__":
    print("Generating sample workbook...")
    filepath = generate_sample_workbook()
    print(f"\nSample workbook created: {filepath}")
    print("\nThis demonstrates the output format. In production, data would be")
    print("populated from the 13F scraper, PubMed KOL extractor, and other sources.")
